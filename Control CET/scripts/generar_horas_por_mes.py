import os
import shutil
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_DIR   = os.path.join(BASE, "Input")
MATRIX_PATH = os.path.join(INPUT_DIR, "Communication Matrix (1).xlsx")
OUTPUT_DIR  = os.path.join(BASE, "Outputs")
BACKUP_DIR  = os.path.join(BASE, "Backups")
FECHA_HOY   = datetime.now().strftime("%Y%m%d")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, f"Horas_NuevaEPS_PorMes_{FECHA_HOY}.xlsx")

PROJECT_FILTER = "CO - Nueva EPS - Migracion ROSA Fixed"


def resolver_horas_path(input_dir):
    """Busca Horas.xlsx/.xls/.csv en Input y devuelve la ruta del más reciente."""
    candidatos = []
    for nombre in os.listdir(input_dir):
        base, ext = os.path.splitext(nombre)
        if base.lower() == "horas" and ext.lower() in (".xlsx", ".xls", ".csv"):
            candidatos.append(os.path.join(input_dir, nombre))
    if not candidatos:
        raise FileNotFoundError(
            f"No se encontró Horas.xlsx, Horas.xls ni Horas.csv en {input_dir}"
        )
    candidatos.sort(key=os.path.getmtime, reverse=True)
    return candidatos[0]


def leer_horas(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        for sep in (",", ";", "\t", "|"):
            try:
                df = pd.read_csv(path, sep=sep, encoding="utf-8-sig")
                if df.shape[1] > 1:
                    return df
            except Exception:
                continue
        return pd.read_csv(path, encoding="utf-8-sig", engine="python", sep=None)
    return pd.read_excel(path)


def normalizar_nombre(s):
    """Devuelve un set de tokens en minúsculas para comparar nombres en cualquier orden."""
    if not isinstance(s, str):
        return frozenset()
    limpio = s.replace(",", " ").replace(".", " ")
    return frozenset(t for t in limpio.lower().split() if t)

MESES_ES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic",
}

# 1. Cargar AllStar
matrix = pd.read_excel(MATRIX_PATH)
matrix.columns = matrix.columns.str.strip()
allstar_emails = set(
    matrix["EMAIL ADDRESS"].dropna().astype(str).str.strip().str.lower()
)
allstar_nombres = {
    normalizar_nombre(n)
    for n in matrix.get("NAME", pd.Series(dtype=str)).dropna().astype(str)
}
allstar_nombres.discard(frozenset())
print(f"AllStar emails cargados: {len(allstar_emails)}")

# 2. Cargar y filtrar archivo de horas (Excel o CSV)
HORAS_PATH = resolver_horas_path(INPUT_DIR)
df = leer_horas(HORAS_PATH)
print(f"Archivo de horas cargado: {HORAS_PATH}")
df.columns = df.columns.str.strip()
df["Date"]    = pd.to_datetime(df["Date"], errors="coerce")
df["Hours"]   = pd.to_numeric(df["Hours"], errors="coerce").fillna(0)

# Identificador de usuario: usa "User ID" si existe (email), si no usa "User" (nombre)
if "User ID" in df.columns:
    df["uid"] = df["User ID"].astype(str).str.strip().str.lower()
    uid_es_email = True
else:
    df["uid"] = df["User"].astype(str).str.strip()
    uid_es_email = False
print(f"Identificador de usuario: {'User ID (email)' if uid_es_email else 'User (nombre)'}")

mask = (df["Project"].str.strip() == PROJECT_FILTER)
f = df[mask].dropna(subset=["Date"]).copy()
f["MesKey"] = f["Date"].dt.to_period("M")
print(f"Filas filtradas para {PROJECT_FILTER}: {len(f)}")

# 3. Pivot horas por usuario y mes
pivot = (
    f.groupby(["uid", "MesKey"])["Hours"].sum()
     .unstack(fill_value=0)
     .sort_index(axis=1)
)

# 4. Nombre y Group por uid
group_por_user = (
    f.groupby("uid")["Group"]
     .agg(lambda s: s.dropna().mode().iloc[0] if not s.dropna().empty else "")
)
nombre_por_user = (
    f.groupby("uid")["User"]
     .agg(lambda s: s.dropna().mode().iloc[0] if not s.dropna().empty else "")
)

meses = list(pivot.columns)
print(f"Meses cubiertos: {[str(m) for m in meses]}")


def es_allstar(uid, nombre):
    if uid_es_email and uid in allstar_emails:
        return True
    return normalizar_nombre(nombre) in allstar_nombres


allstar_detectados = sum(
    1 for uid in pivot.index if es_allstar(uid, nombre_por_user.get(uid, uid))
)
print(f"Usuarios AllStar detectados en datos: {allstar_detectados}")

# 5. Construir filas ordenadas por Nombre
filas = []
for uid, row in pivot.iterrows():
    filas.append({
        "uid":    uid,
        "Nombre": nombre_por_user.get(uid, uid),
        "Group":  group_por_user.get(uid, ""),
        "horas":  row,
    })
filas.sort(key=lambda x: x["Nombre"].lower())

# 6. Escribir Excel
wb = Workbook()
ws = wb.active
ws.title = "Horas Nueva EPS por Mes"

header_font  = Font(bold=True, color="FFFFFF")
header_fill  = PatternFill("solid", fgColor="2F4F8F")
header_align = Alignment(horizontal="center", vertical="center")
total_font   = Font(bold=True, color="FFFFFF")
total_fill   = PatternFill("solid", fgColor="2F4F8F")
fill_blanco  = PatternFill("solid", fgColor="FFFFFF")
fill_gris    = PatternFill("solid", fgColor="F2F2F2")

# Encabezado
headers = ["Nombre", "Group"] + [
    f"{MESES_ES[m.month]} {m.year}" for m in meses
]
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = header_align

# Filas de datos
for row_idx, fila in enumerate(filas, start=2):
    fill = fill_blanco if row_idx % 2 == 0 else fill_gris
    ws.cell(row=row_idx, column=1, value=fila["Nombre"]).fill = fill
    ws.cell(row=row_idx, column=2, value=fila["Group"]).fill  = fill
    is_allstar = es_allstar(fila["uid"], fila["Nombre"])
    for col_offset, mes in enumerate(meses, start=3):
        hrs = fila["horas"][mes]
        if is_allstar and hrs > 0:
            value = f"={hrs}*0.5"
        else:
            value = 0 if hrs == 0 else hrs
        cell = ws.cell(row=row_idx, column=col_offset, value=value)
        cell.fill = fill

# Fila TOTAL
total_row = len(filas) + 2
ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
ws.cell(row=total_row, column=1).fill               = total_fill
ws.cell(row=total_row, column=2, value="").fill      = total_fill
for col_offset in range(3, 3 + len(meses)):
    col_letter = get_column_letter(col_offset)
    formula    = f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
    cell = ws.cell(row=total_row, column=col_offset, value=formula)
    cell.font = total_font
    cell.fill = total_fill

# 7. Ajuste de ancho de columnas
for col_idx, col_cells in enumerate(ws.columns, start=1):
    max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
    min_w   = 12 if col_idx > 2 else 20
    ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, min_w)

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# Mover Excels existentes en Outputs a Backups antes de generar el nuevo
for nombre in os.listdir(OUTPUT_DIR):
    if nombre.lower().endswith((".xlsx", ".xls")):
        origen  = os.path.join(OUTPUT_DIR, nombre)
        destino = os.path.join(BACKUP_DIR, nombre)
        if os.path.exists(destino):
            base, ext = os.path.splitext(nombre)
            destino = os.path.join(
                BACKUP_DIR,
                f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}",
            )
        shutil.move(origen, destino)
        print(f"Backup: {origen} -> {destino}")

wb.save(OUTPUT_PATH)
print(f"Archivo guardado en: {OUTPUT_PATH}")
